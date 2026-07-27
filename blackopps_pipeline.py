#!/usr/bin/env python3
"""
BlackOpps Pipeline — צינור עיבוד מלא: Excel → ייבוא → סיווג GOTV → Battle Plan

Usage:
    python3 blackopps_pipeline.py                          # מעבד את קובץ file.xlsx בתיקיית upload
    python3 blackopps_pipeline.py --file voters_new.xlsx   # מעבד קובץ ספציפי
    python3 blackopps_pipeline.py --skip-import            # קובץ כבר ב-DB, רק סיווג
    python3 blackopps_pipeline.py --dry-run                # מציג מה יקרה בלי לשנות

Outputs:
    gotv_battle_plan.json          — תוכנית קרב מלאה (TOP 200 + סגמנטציית ערוצים)
    gotv_voter_classifications.csv — כל הבוחרים ממוינים לפי עדיפות
"""
import asyncio
import hashlib
import json
import os
import random
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import aiohttp
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent / "backend"))
from app.intelligence.gotv import GOTVPredictor, GOTVProfile, VoterCategory, gotv_battleplan
from app.intelligence.scoring import InfluenceProfile, InfluenceTier

API_BASE = os.getenv("BLACKOPPS_API_URL", "http://127.0.0.1:8000")
DEFAULT_EXCEL = str(Path(__file__).resolve().parent / "upload" / "file.xlsx")
IMPORT_BATCH = 50
PUSH_BATCH = 50
random.seed(42)

# ─── Helpers ───────────────────────────────────────────────────────────

def load_voters(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]

    voters = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        first = str(data.get("שם פרטי", "")).strip()
        last  = str(data.get("שם משפחה", "")).strip()
        if not first or not last:
            continue

        phone   = str(data.get("מס טלפון 1", "")).strip() if data.get("מס טלפון 1") else ""
        street  = str(data.get("רחוב", "")).strip() if data.get("רחוב") else ""
        house   = str(data.get("מס בית", "")).strip() if data.get("מס בית") else ""
        branch  = str(data.get("סניף בתנועה", "")).strip() if data.get("סניף בתנועה") else ""
        founders = str(data.get("מייסדים", "")).strip() if data.get("מייסדים") else ""

        raw = f"{first}:{last}:{street}:{house}"
        national_id = hashlib.sha256(raw.encode()).hexdigest()[:20]

        voters.append({
            "national_id": national_id,
            "first_name": first,
            "last_name": last,
            "full_name": f"{first} {last}",
            "city": str(data.get("ישוב", "פתח תקווה")).strip(),
            "neighborhood": street,
            "phone": phone[:20] if phone and len(phone) > 5 else None,
            "raw_data": {
                "street": street,
                "house": house,
                "branch": branch,
                "founders": founders,
                "source": "ליכוד פנקס בוחרים",
            },
            "has_phone": bool(phone and len(phone) > 5),
            "branch": branch,
            "founders": founders,
        })
    return voters


def make_synthetic_profile(voter: dict) -> InfluenceProfile:
    seed_val = int(hashlib.md5(voter["full_name"].encode()).hexdigest()[:8], 16) % 100
    rng = random.Random(seed_val)

    founders_bonus = 25 if voter.get("founders") and len(voter["founders"]) > 2 else 0
    branch_bonus  = 15 if voter.get("branch") and len(voter["branch"]) > 2 else 0
    phone_bonus   = 18 if voter.get("has_phone") else 0

    political_capital   = min(90, max(8,  25 + founders_bonus + rng.gauss(15, 12)))
    community_influence  = min(90, max(5,  20 + branch_bonus  + rng.gauss(15, 14)))
    voter_reliability    = min(95, max(10, 40 + phone_bonus   + rng.gauss(12, 15)))
    financial_leverage   = min(85, max(2,  rng.gauss(25, 18)))

    composite = (
        political_capital  * 0.30 +
        community_influence * 0.25 +
        voter_reliability   * 0.25 +
        financial_leverage  * 0.20
    )
    tier = (
        InfluenceTier.CRITICAL if composite >= 85 else
        InfluenceTier.HIGH     if composite >= 70 else
        InfluenceTier.MODERATE if composite >= 50 else
        InfluenceTier.LOW      if composite >= 30 else
        InfluenceTier.NEGLIGIBLE
    )
    return InfluenceProfile(
        name=voter["full_name"],
        political_capital=round(political_capital, 1),
        community_influence=round(community_influence, 1),
        voter_reliability=round(voter_reliability, 1),
        financial_leverage=round(financial_leverage, 1),
        composite_score=round(composite, 1),
        tier=tier,
    )


def make_voting_history(profile: InfluenceProfile) -> dict:
    vr = profile.voter_reliability
    if vr >= 80:    consistency = "always"
    elif vr >= 55:  consistency = "usually"
    elif vr >= 35:  consistency = "sometimes"
    elif vr >= 15:  consistency = "rarely"
    else:           consistency = "never"

    elections = []
    for i in range(5):
        voted = (
            i < (5 if consistency == "always" else 4 if consistency == "usually"
                 else 3 if consistency == "sometimes" else 1 if consistency == "rarely" else 0)
        )
        elections.append({"election": f"202{2-i}", "voted": voted})
    years = {"always": 15, "usually": 10, "sometimes": 5, "rarely": 3, "never": 1}
    return {"consistency": consistency, "years_registered": years.get(consistency, 5), "recent_elections": elections}

# ─── Phase 1: Import ────────────────────────────────────────────────────

async def import_voters(session: aiohttp.ClientSession, voters: list[dict], existing_ids: set) -> tuple[int, int]:
    new_voters = [v for v in voters if v["national_id"] not in existing_ids]
    if not new_voters:
        print("✅ כל הבוחרים כבר קיימים ב-DB — מדלג על ייבוא", flush=True)
        return 0, 0

    print(f"📥 מייבא {len(new_voters)} בוחרים חדשים ({len(voters) - len(new_voters)} כבר קיימים)...", flush=True)
    success, fail = 0, 0

    for i in range(0, len(new_voters), IMPORT_BATCH):
        batch = new_voters[i : i + IMPORT_BATCH]
        tasks = [
            session.post(f"{API_BASE}/voters", json={
                "national_id": v["national_id"],
                "first_name": v["first_name"],
                "last_name": v["last_name"],
                "city": v["city"],
                "neighborhood": v["neighborhood"],
                "phone": v.get("phone"),
                "raw_data": v.get("raw_data"),
            }, timeout=aiohttp.ClientTimeout(total=15))
            for v in batch
        ]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for r in results:
            if isinstance(r, Exception):
                fail += 1
            elif r.status in (200, 201):
                success += 1
            else:
                fail += 1

        pct = min(100, (i + len(batch)) * 100 // len(new_voters))
        print(f"  ייבוא: {i+len(batch)}/{len(new_voters)} ({pct}%) — ✓{success} ✗{fail}", flush=True)

    print(f"📥 ייבוא הושלם: ✓{success} הצליחו, ✗{fail} נכשלו", flush=True)
    return success, fail

# ─── Phase 2: GOTV Classification ────────────────────────────────────────

def classify_voters(voters: list[dict]) -> list[GOTVProfile]:
    print(f"🔧 מייצר פרופילי השפעה ל-{len(voters)} בוחרים...", flush=True)
    profiles = {}
    histories = {}
    for v in voters:
        p = make_synthetic_profile(v)
        profiles[v["full_name"]] = p
        histories[v["full_name"]] = make_voting_history(p)

    print(f"🗳  מריץ GOTV Predictor (local)...", flush=True)
    predictor = GOTVPredictor()
    t0 = time.time()
    results = [predictor.predict(name, profiles[name], histories.get(name, {})) for name in profiles]
    print(f"⏱  {len(results)} בוחרים סווגו ב-{time.time() - t0:.1f}s", flush=True)
    return results

# ─── Phase 3: Push scores ────────────────────────────────────────────────

async def get_existing_ids(session: aiohttp.ClientSession) -> set:
    try:
        async with session.get(f"{API_BASE}/voters?limit=5000", timeout=aiohttp.ClientTimeout(total=45)) as resp:
            data = await resp.json()
            return {v["national_id"] for v in data}
    except Exception:
        return set()

async def push_scores(session: aiohttp.ClientSession, results: list[GOTVProfile],
                      name_to_id: dict) -> tuple[int, int]:
    """
    Push GOTV classification metadata to Railway.
    Note: turnout_score and support_score are read-only in the API schema.
    We push category + priority into raw_data for dashboard visibility.
    The authoritative GOTV data lives in the battle plan JSON/CSV files.
    """
    print(f"☁️  דוחף מטה-דאטה GOTV ל-Railway ({len(results)} בוחרים)...", flush=True)
    success, fail = 0, 0
    prepared = []

    for p in results:
        vid = name_to_id.get(p.name)
        if vid:
            prepared.append({
                "voter_id": vid,
                "raw_data": {
                    "gotv_category": p.category.value,
                    "gotv_priority": round(p.priority_score, 1),
                    "gotv_turnout": round(p.turnout_probability, 1),
                    "gotv_persuasion": round(p.persuasion_score, 1),
                    "gotv_channel": p.optimal_channel.value,
                    "gotv_frequency": p.contact_frequency,
                    "gotv_frame": p.messaging_frame,
                    "classified_at": datetime.now(timezone.utc).isoformat(),
                }
            })

    for i in range(0, len(prepared), PUSH_BATCH):
        batch = prepared[i : i + PUSH_BATCH]
        tasks = [
            session.patch(f"{API_BASE}/voters/{b['voter_id']}", json={
                "raw_data": b["raw_data"]
            }, timeout=aiohttp.ClientTimeout(total=10))
            for b in batch
        ]
        g_results = await asyncio.gather(*tasks, return_exceptions=True)
        for r in g_results:
            if isinstance(r, Exception) or r.status not in (200, 201, 204):
                fail += 1
            else:
                success += 1

        pct = min(100, (i + len(batch)) * 100 // len(prepared))
        print(f"  PUSH: {i+len(batch)}/{len(prepared)} ({pct}%) — ✓{success} ✗{fail}", flush=True)

    print(f"☁️  מטה-דאטה נדחף: ✓{success} ✗{fail}", flush=True)
    return success, fail

# ─── Phase 4: Generate Battle Plan ───────────────────────────────────────

def generate_output(results: list[GOTVProfile], voters: list[dict], output_dir: str = "."):
    cats = Counter(p.category.value for p in results)
    total = len(results)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    # Top 200
    top_200 = sorted(
        [{
            "name": p.name,
            "category": p.category.value,
            "priority": round(p.priority_score, 1),
            "turnout": round(p.turnout_probability, 1),
            "persuasion": round(p.persuasion_score, 1),
            "channel": p.optimal_channel.value,
            "frequency": p.contact_frequency,
            "frame": p.messaging_frame,
            "dropout_risk": round(p.dropout_risk, 1),
            "volunteer": p.volunteer_potential,
            "donor": p.donor_potential,
            "multiplier": p.multiplier_potential,
            "action": p.recommended_action,
        } for p in results],
        key=lambda x: x["priority"], reverse=True
    )[:200]

    # Channel segmentation
    channels = {}
    for p in results:
        ch = p.optimal_channel.value
        channels.setdefault(ch, []).append(p.name)

    bp = gotv_battleplan(results)

    battle_plan = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_file": Path(DEFAULT_EXCEL).name,
        "total_voters": total,
        "classification": {
            "safe": cats.get("safe", 0),
            "leaning": cats.get("leaning", 0),
            "swing": cats.get("swing", 0),
            "at_risk": cats.get("at_risk", 0),
            "lost": cats.get("lost", 0),
        },
        "resource_allocation": bp.get("resource_allocation", {}),
        "top_10_priority": bp.get("top_10_priority", []),
        "top_200_priority": top_200,
        "field_ops_by_channel": {ch: {"count": len(n), "voters": n[:30]} for ch, n in channels.items()},
        "projected_field_ops_needed": max(1, total // 200),
    }

    bp_path = f"{output_dir}/gotv_battle_plan_{ts}.json"
    with open(bp_path, "w", encoding="utf-8") as f:
        json.dump(battle_plan, f, ensure_ascii=False, indent=2)
    print(f"📁 Battle Plan: {bp_path}", flush=True)

    csv_path = f"{output_dir}/gotv_voter_classifications_{ts}.csv"
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("שם,קטגוריה,עדיפות,הסתברות הצבעה,שכנוע,ערוץ,תדירות,מסר,סיכון נשירה,מתנדב,תורם,מכפיל\n")
        for p in sorted(results, key=lambda x: x.priority_score, reverse=True):
            f.write(f'"{p.name}",{p.category.value},{p.priority_score:.0f},{p.turnout_probability:.0f},{p.persuasion_score:.0f},'
                    f'{p.optimal_channel.value},{p.contact_frequency},{p.messaging_frame},{p.dropout_risk:.0f},'
                    f'{p.volunteer_potential},{p.donor_potential},{p.multiplier_potential}\n')
    print(f"📁 CSV: {csv_path}", flush=True)

    # Also save latest copies for quick access
    for src, dst in [(bp_path, f"{output_dir}/gotv_battle_plan.json"),
                     (csv_path, f"{output_dir}/gotv_voter_classifications.csv")]:
        with open(src, "r") as fsrc, open(dst, "w") as fdst:
            fdst.write(fsrc.read())

    return battle_plan

# ─── Main ────────────────────────────────────────────────────────────────

async def main():
    skip_import = "--skip-import" in sys.argv
    dry_run    = "--dry-run" in sys.argv
    filepath   = DEFAULT_EXCEL

    for a in sys.argv[1:]:
        if a.startswith("--file="):
            filepath = a.split("=", 1)[1]

    if not Path(filepath).exists():
        print(f"❌ קובץ לא נמצא: {filepath}")
        return

    print("=" * 65)
    print("  BLACKOPPS PIPELINE — Excel → GOTV → Battle Plan 🔴")
    print(f"  API:  {API_BASE}")
    print(f"  קובץ: {filepath}")
    print(f"  ייבוא: {'לא' if skip_import else 'כן'}  |  Dry run: {'כן' if dry_run else 'לא'}")
    print("=" * 65, flush=True)

    # ── Load ──
    voters = load_voters(filepath)
    print(f"\n📋 {len(voters)} בוחרים נטענו", flush=True)
    if not voters:
        return

    if dry_run:
        for i, v in enumerate(voters[:10]):
            print(f"  {i+1}. {v['first_name']} {v['last_name']} | {v['city']} | {v.get('neighborhood','')} | tel={v.get('phone','-')}")
        print(f"  ... ועוד {len(voters)-10}")
        print("🔍 Dry run — בוצע. לא נגע ב-API.")
        return

    connector = aiohttp.TCPConnector(limit=10)
    async with aiohttp.ClientSession(connector=connector) as session:
        # ── Phase 1: Import ──
        if not skip_import:
            existing_ids = await get_existing_ids(session)
            print(f"  קיימים ב-DB: {len(existing_ids)}", flush=True)
            await import_voters(session, voters, existing_ids)

        # Re-fetch IDs after import (some may be new)
        existing_ids = await get_existing_ids(session)
        print(f"  סה\"כ ב-DB אחרי ייבוא: {len(existing_ids)}", flush=True)

        # ── Phase 2: Classify locally ──
        results = classify_voters(voters)

        # ── Phase 3: Push scores ──
        async with session.get(f"{API_BASE}/voters?limit=5000", timeout=aiohttp.ClientTimeout(total=45)) as resp:
            api_voters = await resp.json()
        name_to_id = {f"{v['first_name']} {v['last_name']}": v["id"] for v in api_voters}
        await push_scores(session, results, name_to_id)

    # ── Phase 4: Generate outputs ──
    bp = generate_output(results, voters)

    # ── Summary ──
    c = bp["classification"]
    print("\n" + "=" * 65)
    print("  🎯 תוצאות GOTV")
    print("=" * 65)
    print(f"  🟢 SAFE:     {c['safe']:5d} ({c['safe']*100/bp['total_voters']:.1f}%)")
    print(f"  🔵 LEANING:  {c['leaning']:5d} ({c['leaning']*100/bp['total_voters']:.1f}%)")
    print(f"  🟡 SWING:    {c['swing']:5d} ({c['swing']*100/bp['total_voters']:.1f}%)  ⚡")
    print(f"  🔴 AT_RISK:  {c['at_risk']:5d} ({c['at_risk']*100/bp['total_voters']:.1f}%)")
    print(f"  ⚫ LOST:     {c['lost']:5d} ({c['lost']*100/bp['total_voters']:.1f}%)")
    print(f"  פעילי שטח:   {bp['projected_field_ops_needed']}")
    print(f"\n  🔝 TOP 5:")
    for i, v in enumerate(bp["top_10_priority"][:5]):
        print(f"  {i+1}. {v['name']:30s} [{v['category']}] pri={v['priority']}")
    print("=" * 65)
    print("\n✅ צינור הושלם. קבצי שטח מוכנים.", flush=True)


if __name__ == "__main__":
    asyncio.run(main())
